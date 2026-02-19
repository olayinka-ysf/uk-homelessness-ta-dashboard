import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PROCESSED = Path("data/processed")
OUT = Path("data/powerbi")
OUT.mkdir(exist_ok=True)

ts = pd.read_csv(PROCESSED / "timeseries_clean.csv")
la = pd.read_csv(PROCESSED / "la_detailed_clean.csv")
imd = pd.read_csv(PROCESSED / "imd_clean.csv")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ALT_FILL   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD_FONT  = Font(bold=True, size=10)
STD_FONT   = Font(size=10)
HDR_ALIGN  = Alignment(horizontal="left", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BTM   = Border(bottom=Side(style="thin", color="C9C9C9"))

def _autowidth(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 52)

def _write_header(ws, headers):
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN

def _write_rows(ws, df, num_fmts=None):
    num_fmts = num_fmts or {}
    for r, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = STD_FONT
            cell.alignment = CELL_ALIGN
            if c in num_fmts:
                cell.number_format = num_fmts[c]

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Sheet 1: England Summary ──────────────────────────────────────────────────
ws1 = wb.create_sheet("England Summary")
latest = ts.iloc[-1]
prev_yr = ts[(ts["year"] == latest["year"] - 1) & (ts["quarter"] == latest["quarter"])]
yoy_ta = int(latest["total_ta_households"] - prev_yr["total_ta_households"].values[0]) if len(prev_yr) else None
yoy_pct = round(yoy_ta / prev_yr["total_ta_households"].values[0] * 100, 1) if yoy_ta is not None else None

summary_data = [
    ("Metric", "Value", "Quarter"),
    ("Total households in temporary accommodation",    int(latest["total_ta_households"]),  f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Households in TA with dependent children",       int(latest["ta_with_children"]),      f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Children in temporary accommodation",            int(latest["children_in_ta"]),         f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Prevention duties accepted",                     int(latest["prevention_duties"]),      f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Relief duties accepted",                         int(latest["relief_duties"]),           f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Year-on-year change in TA households",           yoy_ta,                                "vs same quarter prior year"),
    ("YoY change (%)",                                 yoy_pct,                               "vs same quarter prior year"),
    ("B&B accommodation households",                   int(latest["bb_accommodation"]),       f"Q{int(latest['quarter'])} {int(latest['year'])}"),
    ("Out of area placements",                         int(latest["out_of_area"]),             f"Q{int(latest['quarter'])} {int(latest['year'])}"),
]

ws1.freeze_panes = "A2"
hdr_row = summary_data[0]
_write_header(ws1, list(hdr_row))
for r, row in enumerate(summary_data[1:], 2):
    fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = STD_FONT
        cell.alignment = CELL_ALIGN
        if c == 2 and isinstance(val, (int, float)):
            cell.number_format = "#,##0" if isinstance(val, int) else "#,##0.0"
_autowidth(ws1)

# ── Sheet 2: Duties Trend ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Duties Trend")
duties_df = ts[ts["prevention_duties"].notna()].copy()
duties_df["yoy_prevention"] = duties_df["prevention_duties"].diff(4)
duties_df["yoy_relief"] = duties_df["relief_duties"].diff(4)
duties_out = duties_df[["year", "quarter", "prevention_duties", "relief_duties", "yoy_prevention", "yoy_relief"]].copy()
duties_out.columns = ["Year", "Quarter", "Prevention Duties", "Relief Duties", "YoY Prevention Change", "YoY Relief Change"]
duties_out = duties_out.reset_index(drop=True)

ws2.freeze_panes = "A2"
_write_header(ws2, list(duties_out.columns))
_write_rows(ws2, duties_out, num_fmts={3: "#,##0", 4: "#,##0", 5: "+#,##0;-#,##0;0", 6: "+#,##0;-#,##0;0"})
_autowidth(ws2)

# ── Sheet 3: TA by Type ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("TA by Type")
ta_type_df = ts[ts["bb_accommodation"].notna()].copy()
ta_type_df["self_contained"] = ta_type_df["leased"].fillna(0) + ta_type_df["la_ha_stock"].fillna(0)
ta_type_df["bb_pct"]         = (ta_type_df["bb_accommodation"] / ta_type_df["total_ta_households"]).round(4)
ta_type_df["nightly_pct"]    = (ta_type_df["nightly_paid"].fillna(0) / ta_type_df["total_ta_households"]).round(4)
ta_type_df["sc_pct"]         = (ta_type_df["self_contained"] / ta_type_df["total_ta_households"]).round(4)
ta_type_out = ta_type_df[["year", "quarter", "total_ta_households", "ta_with_children",
                            "bb_accommodation", "bb_pct", "nightly_paid", "nightly_pct",
                            "self_contained", "sc_pct"]].copy()
ta_type_out.columns = ["Year", "Quarter", "Total TA Households", "TA with Children",
                        "B&B", "B&B %", "Nightly Paid", "Nightly Paid %", "Self-Contained", "Self-Contained %"]
ta_type_out = ta_type_out.reset_index(drop=True)

ws3.freeze_panes = "A2"
_write_header(ws3, list(ta_type_out.columns))
_write_rows(ws3, ta_type_out, num_fmts={3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "0.0%", 7: "#,##0", 8: "0.0%", 9: "#,##0", 10: "0.0%"})
_autowidth(ws3)

# ── Sheet 4: LA Benchmarking ──────────────────────────────────────────────────
ws4 = wb.create_sheet("LA Benchmarking")
la_bm = la[la["households_in_area_000s"] > 0].copy()
la_bm["ta_rate_per_1000"] = (la_bm["total_ta_households"] / la_bm["households_in_area_000s"]).round(2)
la_bm["duties_rate_per_1000"] = ((la_bm["prevention_duties"] + la_bm["relief_duties"]) / la_bm["households_in_area_000s"]).round(2)
la_bm = la_bm.merge(imd[["local_authority_code", "imd_score", "imd_rank", "imd_decile"]], on="local_authority_code", how="left")
la_bm = la_bm.sort_values("ta_rate_per_1000", ascending=False).reset_index(drop=True)
la_bm["ta_rank"] = la_bm["ta_rate_per_1000"].rank(ascending=False, method="min").astype(int)

la_out = la_bm[["local_authority_code", "local_authority_name", "total_ta_households",
                  "ta_rate_per_1000", "prevention_duties", "relief_duties",
                  "duties_rate_per_1000", "imd_score", "imd_rank", "imd_decile", "ta_rank"]].copy()
la_out.columns = ["LA Code", "Local Authority", "Total TA Households", "TA Rate per 1,000",
                   "Prevention Duties", "Relief Duties", "Duties Rate per 1,000",
                   "IMD Score", "IMD Rank", "IMD Decile", "TA Rank"]

ws4.freeze_panes = "A2"
_write_header(ws4, list(la_out.columns))
_write_rows(ws4, la_out, num_fmts={3: "#,##0", 4: "#,##0.00", 5: "#,##0", 6: "#,##0", 7: "#,##0.00", 8: "0.000", 9: "#,##0", 10: "0", 11: "0"})
_autowidth(ws4)

wb.save(OUT / "homelessness_summary.xlsx")
print(f"homelessness_summary.xlsx saved with {len(la_out)} LA rows")
